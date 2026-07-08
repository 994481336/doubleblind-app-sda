from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook


B737_TEMPLATE = "双盲测试表"
GO_AROUND_TEMPLATE = "复飞专项检查表"
UNKNOWN_TEMPLATE = "未知格式"

EXPECTED_DEDUCTION_COLS = {
    B737_TEMPLATE: 93,
    GO_AROUND_TEMPLATE: 149,
}

META_ALIASES = {
    "序号": ["序号"],
    "姓名": ["姓名"],
    "日期": ["日期"],
    "所属单位": ["所属单位", "所属/单位", "所属 单位", "单位"],
    "技术等级": ["技术等级", "职务"],
    "总飞行时间": ["总飞行时间"],
    "本机型经历时间": ["本机型经历时间"],
    "评估员": ["评估员", "检查员"],
    "得分": ["得分", "总得分"],
}


@dataclass
class ParsedWorkbook:
    summary: dict[str, Any]
    ratings: pd.DataFrame
    summary_scores: pd.DataFrame
    deductions: pd.DataFrame
    deduction_columns: pd.DataFrame


def compact_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def key_text(value: Any) -> str:
    return re.sub(r"[\s/／:：()（）\-—_]+", "", compact_text(value)).lower()


def to_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, str):
        value = value.replace("−", "-").replace("－", "-").strip()
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def clean_name(value: Any) -> str:
    text = compact_text(value)
    text = re.sub(r"/.*$", "", text).strip()
    text = re.sub(r"（.*?）", "", text).strip()
    return text


def normalize_role(value: Any) -> str:
    text = compact_text(value)
    if "副" in text:
        return "副驾驶"
    if "机长" in text or "教员" in text or "M" in text.upper():
        return "机长"
    return text or "未识别"


def normalize_unit(value: Any) -> str:
    text = compact_text(value)
    text = re.sub(r"(B737|A320|A330|A350|B777|B787|C919|C909|ARJ21|ARJ)", "", text, flags=re.I)
    text = re.sub(r"复飞专项检查数据采集表", "", text)
    text = re.sub(r"[()（）【】\[\]_-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if "圆通" in text:
        return "圆通航空"
    return text


def find_score_row(ws) -> int:
    for row_idx in range(1, min(ws.max_row, 8) + 1):
        if "分值" in key_text(ws.cell(row_idx, 1).value):
            return row_idx
    return 3


def find_meta_columns(ws, score_row: int) -> dict[str, int]:
    found: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        raw_text = " ".join(
            compact_text(ws.cell(row_idx, col_idx).value)
            for row_idx in range(1, score_row)
            if compact_text(ws.cell(row_idx, col_idx).value)
        )
        normalized = key_text(raw_text)
        if not normalized:
            continue
        for canonical, aliases in META_ALIASES.items():
            if canonical in found:
                continue
            if any(key_text(alias) in normalized for alias in aliases):
                found[canonical] = col_idx
    return found


def get_filled_headers(ws, score_row: int) -> dict[int, list[str]]:
    current_by_row = {row_idx: "" for row_idx in range(1, score_row)}
    headers: dict[int, list[str]] = {}
    for col_idx in range(1, ws.max_column + 1):
        parts: list[str] = []
        for row_idx in range(1, score_row):
            value = compact_text(ws.cell(row_idx, col_idx).value)
            if value:
                current_by_row[row_idx] = value
            inherited = current_by_row[row_idx]
            if inherited and inherited not in parts:
                parts.append(inherited)
        headers[col_idx] = parts
    return headers


def find_deduction_columns(ws, score_row: int) -> list[int]:
    cols: list[int] = []
    for col_idx in range(1, ws.max_column + 1):
        score = to_number(ws.cell(score_row, col_idx).value)
        if score is not None and score < 0:
            cols.append(col_idx)
    return cols


def detect_template(score_row: int, deduction_count: int, headers: dict[int, list[str]]) -> str:
    first_headers = " ".join(" ".join(parts) for col, parts in headers.items() if col <= 20)
    if score_row == 5 or deduction_count > 100 or "复飞" in first_headers:
        return GO_AROUND_TEMPLATE
    if deduction_count == EXPECTED_DEDUCTION_COLS[B737_TEMPLATE]:
        return B737_TEMPLATE
    return UNKNOWN_TEMPLATE


def describe_deduction_column(
    template_type: str,
    col_idx: int,
    score: float,
    headers: dict[int, list[str]],
) -> dict[str, Any]:
    parts = [p for p in headers.get(col_idx, []) if p and key_text(p) not in {"分值"}]
    if template_type == GO_AROUND_TEMPLATE and parts and key_text(parts[0]).startswith("科目"):
        subject = parts[0]
        middle = parts[1:-1]
        standard = parts[-1] if len(parts) > 1 else parts[0]
        scoring_item = " / ".join(middle) if middle else standard
    else:
        subject = parts[0] if parts else "常规评估"
        scoring_item = parts[0] if parts else f"第{col_idx}列"
        standard = parts[-1] if len(parts) > 1 else scoring_item

    display = " / ".join(parts) if parts else f"第{col_idx}列"
    return {
        "模板类型": template_type,
        "原始列号": col_idx,
        "科目名称": subject,
        "评分项目": scoring_item,
        "扣分标准": standard,
        "标准分值": score,
        "扣分项": display,
    }


def source_unit_from_file(file_name: str) -> str:
    stem = re.sub(r"\.(xlsx|xls)$", "", file_name, flags=re.I)
    text = normalize_unit(stem)
    return compact_text(text or stem)


def row_has_any(ws, row_idx: int, cols: list[int]) -> bool:
    return any(ws.cell(row_idx, col_idx).value not in (None, "") for col_idx in cols)


def parse_workbook(file_bytes: bytes, file_name: str) -> ParsedWorkbook:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    score_row = find_score_row(ws)
    meta_cols = find_meta_columns(ws, score_row)
    headers = get_filled_headers(ws, score_row)
    deduction_col_indexes = find_deduction_columns(ws, score_row)
    template_type = detect_template(score_row, len(deduction_col_indexes), headers)
    warnings: list[str] = []

    expected_cols = EXPECTED_DEDUCTION_COLS.get(template_type)
    if expected_cols and len(deduction_col_indexes) != expected_cols:
        warnings.append(f"扣分列数应为 {expected_cols}，实际识别 {len(deduction_col_indexes)}。")
    if template_type == UNKNOWN_TEMPLATE:
        warnings.append("未匹配固定格式，仅按负分列尝试解析。")
    if "评估员" not in meta_cols:
        warnings.append("未识别检查员/评估员列。")

    deduction_meta = [
        describe_deduction_column(
            template_type,
            col_idx,
            float(to_number(ws.cell(score_row, col_idx).value) or 0),
            headers,
        )
        for col_idx in deduction_col_indexes
    ]
    deduction_columns = pd.DataFrame(deduction_meta)

    fill_fields = ["序号", "姓名", "日期", "所属单位", "技术等级", "总飞行时间", "本机型经历时间"]
    last_values = {field: "" for field in fill_fields}
    ratings: list[dict[str, Any]] = []
    summary_scores: list[dict[str, Any]] = []
    deduction_events: list[dict[str, Any]] = []
    seen_people: set[str] = set()
    source_unit = source_unit_from_file(file_name)

    row_check_cols = list(meta_cols.values()) + deduction_col_indexes
    for row_idx in range(score_row + 1, ws.max_row + 1):
        if not row_has_any(ws, row_idx, row_check_cols):
            continue
        if compact_text(ws.cell(row_idx, meta_cols.get("日期", 3)).value).startswith("结论"):
            continue

        raw_values: dict[str, Any] = {}
        for field, col_idx in meta_cols.items():
            raw_values[field] = ws.cell(row_idx, col_idx).value

        for field in fill_fields:
            value = raw_values.get(field)
            if value not in (None, ""):
                last_values[field] = value
            raw_values[field] = last_values.get(field, "")

        if not raw_values.get("姓名"):
            continue

        seq = compact_text(raw_values.get("序号"))
        name = clean_name(raw_values.get("姓名"))
        unit = normalize_unit(raw_values.get("所属单位")) or source_unit
        role = normalize_role(raw_values.get("技术等级"))
        evaluator = compact_text(raw_values.get("评估员"))
        template_score = to_number(raw_values.get("得分"))
        person_id = f"{file_name}::{seq}::{name}"
        seen_people.add(person_id)

        if not evaluator:
            continue
        if re.search(r"合计|平均", evaluator):
            summary_scores.append(
                {
                    "人员ID": person_id,
                    "姓名": name,
                    "所属单位": unit,
                    "技术等级": role,
                    "评估员": evaluator,
                    "模板得分": template_score,
                    "来源文件": file_name,
                    "模板类型": template_type,
                }
            )
            continue

        deduction_values: dict[int, float] = {}
        for col_idx in deduction_col_indexes:
            value = to_number(ws.cell(row_idx, col_idx).value)
            if value is not None and value < 0:
                deduction_values[col_idx] = float(value)

        deduction_total = sum(deduction_values.values())
        computed_score = 100 + deduction_total
        rating = {
            "人员ID": person_id,
            "序号": seq,
            "姓名": name,
            "日期": raw_values.get("日期"),
            "所属单位": unit,
            "技术等级": role,
            "总飞行时间": raw_values.get("总飞行时间"),
            "本机型经历时间": raw_values.get("本机型经历时间"),
            "评估员": evaluator,
            "模板得分": template_score,
            "计算得分": computed_score,
            "最终得分": template_score if template_score is not None else computed_score,
            "扣分总和": deduction_total,
            "失分": abs(deduction_total),
            "扣分项数量": len(deduction_values),
            "来源文件": file_name,
            "模板类型": template_type,
        }
        ratings.append(rating)

        for meta in deduction_meta:
            col_idx = int(meta["原始列号"])
            if col_idx not in deduction_values:
                continue
            value = deduction_values[col_idx]
            deduction_events.append(
                {
                    "人员ID": person_id,
                    "姓名": name,
                    "所属单位": unit,
                    "技术等级": role,
                    "评估员": evaluator,
                    "科目名称": meta["科目名称"],
                    "评分项目": meta["评分项目"],
                    "扣分标准": meta["扣分标准"],
                    "扣分项": meta["扣分项"],
                    "标准分值": meta["标准分值"],
                    "扣分值": value,
                    "失分": abs(value),
                    "原始列号": col_idx,
                    "来源文件": file_name,
                    "模板类型": template_type,
                }
            )

    if not ratings:
        warnings.append("未识别到有效评分行。")

    summary = {
        "文件名": file_name,
        "模板类型": template_type,
        "工作表": ws.title,
        "分值行": score_row,
        "人员数": len(seen_people),
        "评分行数": len(ratings),
        "平均/合计行数": len(summary_scores),
        "扣分列数": len(deduction_col_indexes),
        "警告": "；".join(warnings) if warnings else "",
    }
    return ParsedWorkbook(
        summary=summary,
        ratings=pd.DataFrame(ratings),
        summary_scores=pd.DataFrame(summary_scores),
        deductions=pd.DataFrame(deduction_events),
        deduction_columns=deduction_columns,
    )


def parse_many(files: list[tuple[str, bytes]]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    summaries: list[dict[str, Any]] = []
    ratings: list[pd.DataFrame] = []
    summary_scores: list[pd.DataFrame] = []
    deductions: list[pd.DataFrame] = []

    for file_name, file_bytes in files:
        try:
            parsed = parse_workbook(file_bytes, file_name)
        except Exception as exc:  # pragma: no cover - surfaced in Streamlit and tests.
            summaries.append(
                {
                    "文件名": file_name,
                    "模板类型": UNKNOWN_TEMPLATE,
                    "工作表": "",
                    "分值行": None,
                    "人员数": 0,
                    "评分行数": 0,
                    "平均/合计行数": 0,
                    "扣分列数": 0,
                    "警告": f"解析失败：{exc}",
                }
            )
            continue
        summaries.append(parsed.summary)
        ratings.append(parsed.ratings)
        summary_scores.append(parsed.summary_scores)
        deductions.append(parsed.deductions)

    return (
        pd.DataFrame(summaries),
        pd.concat(ratings, ignore_index=True) if ratings else pd.DataFrame(),
        pd.concat(summary_scores, ignore_index=True) if summary_scores else pd.DataFrame(),
        pd.concat(deductions, ignore_index=True) if deductions else pd.DataFrame(),
    )
